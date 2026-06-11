"""
utils/excel_handler.py — Reads Sheet1 (contacts) and Sheet2 (templates).
Writes Remark and Email Status back to Sheet1 after each send.
"""

import os
from openpyxl import load_workbook
from config import Config


class ExcelHandler:

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb       = None
        self.ws1      = None   # Sheet1: contacts
        self.ws2      = None   # Sheet2: templates

    def load(self):
        """Load workbook. Raises FileNotFoundError if path is wrong."""
        self.wb  = load_workbook(self.filepath)

        if "Sheet1" not in self.wb.sheetnames:
            raise ValueError(f"'Sheet1' not found in {self.filepath}. Sheets: {self.wb.sheetnames}")
        if "Sheet2" not in self.wb.sheetnames:
            raise ValueError(f"'Sheet2' not found in {self.filepath}. Sheets: {self.wb.sheetnames}")

        self.ws1 = self.wb["Sheet1"]
        self.ws2 = self.wb["Sheet2"]

    def save(self):
        """Save workbook back to the same file."""
        if self.wb:
            self.wb.save(self.filepath)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet2: template reader
    # ─────────────────────────────────────────────────────────────────────────

    def get_template(self, template_index: int = 0) -> tuple[str, str]:
        """
        Returns (subject_line, body_template, image_path) from Sheet2.
        template_index=0 → first template (row 2, since row 1 is header).
        image_path is None if column D is blank.
        Returns (None, None, None) if not found.
        """
        excel_row = template_index + 2   # row 2 = first data row

        try:
            subject    = self.ws2.cell(row=excel_row, column=2).value  # Column B
            template   = self.ws2.cell(row=excel_row, column=3).value  # Column C
            image_val  = self.ws2.cell(row=excel_row, column=4).value  # Column D — Image Path

            # Resolve image path — support both relative and absolute paths
            image_path = None
            if image_val:
                image_val = str(image_val).strip()
                if os.path.isabs(image_val):
                    image_path = image_val
                else:
                    # Relative path — resolve from project BASE_DIR
                    image_path = os.path.join(Config.BASE_DIR, image_val)

            return str(subject or "").strip(), str(template or "").strip(), image_path
        except Exception:
            return None, None, None

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet1: contact row reader
    # ─────────────────────────────────────────────────────────────────────────

    def get_pending_rows(self) -> list:
        """
        Return all rows in Sheet1 that:
        - Remark column is NOT 'Done'
        - Email Status is blank (not already marked invalid/not found)
        """
        pending = []

        for row in self.ws1.iter_rows(min_row=2):
            # Skip completely empty rows
            if not any(cell.value for cell in row):
                continue

            remark_val = str(row[Config.COL_REMARK].value or "").strip().lower()
            status_val = str(row[Config.COL_STATUS].value or "").strip().lower()

            # Skip already done
            if remark_val == Config.REMARK_DONE.lower():
                continue

            # Skip already marked as errored (optional — remove if you want retry on errors)
            if status_val in (Config.STATUS_INVALID.lower(), Config.STATUS_NOT_FOUND.lower()):
                continue

            pending.append(row)

        return pending

    # ─────────────────────────────────────────────────────────────────────────
    # Reporting helper
    # ─────────────────────────────────────────────────────────────────────────

    def get_summary(self) -> dict:
        """Return a summary dict of statuses in Sheet1."""
        total = done = invalid = not_found = pending = 0

        for row in self.ws1.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            total += 1
            remark = str(row[Config.COL_REMARK] or "").strip().lower()
            status = str(row[Config.COL_STATUS] or "").strip().lower()

            if remark == Config.REMARK_DONE.lower():
                done += 1
            elif status == Config.STATUS_INVALID.lower():
                invalid += 1
            elif status == Config.STATUS_NOT_FOUND.lower():
                not_found += 1
            else:
                pending += 1

        return {
            "total": total,
            "done": done,
            "invalid": invalid,
            "not_found": not_found,
            "pending": pending
        }
