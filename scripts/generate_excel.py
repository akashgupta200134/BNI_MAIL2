"""
scripts/generate_excel.py — Creates a fresh email_campaign.xlsx in data/.
Run this once to generate the template file, then fill in your contacts.

Usage: python scripts/generate_excel.py
"""

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import Config


def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, bg_hex: str):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill      = PatternFill("solid", start_color=bg_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = make_border()


def style_data(cell, alt: bool = False):
    cell.font      = Font(name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color="F2F7FC" if alt else "FFFFFF")
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border    = make_border()


def build_sheet1(wb):
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["Name", "Email", "Email Status", "Remark"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h), "2E6DA4")

    sample_rows = [
        ("Alice Johnson",  "alice.johnson@example.com",  "",              ""),
        ("Bob Smith",      "bob.smith@example.com",      "",              ""),
        ("Carol White",    "carol.white@example.com",    "",              ""),
        ("David Brown",    "david.brown@example.com",    "",              "Done"),
        ("Eve Davis",      "eve-invalid-email",          "Invalid Email", ""),
        ("Frank Miller",   "frank.miller@example.com",   "",              ""),
        ("Grace Lee",      "grace.lee@example.com",      "",              ""),
        ("Henry Wilson",   "henry.wilson@example.com",   "",              ""),
    ]

    for r, (name, email, status, remark) in enumerate(sample_rows, 2):
        alt = (r % 2 == 0)
        for c, val in enumerate([name, email, status, remark], 1):
            style_data(ws.cell(row=r, column=c, value=val), alt)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def build_sheet2(wb):
    ws = wb.create_sheet("Sheet2")
    headers = ["Template No", "Subject Line", "Email Template Body", "Image Path"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h), "2E7D32")

    templates = [
        (
            "Template 1",
            "Exciting Opportunity Just for You – Let's Connect!",
            "Dear {{Name}},\n\nI hope this message finds you well.\n\n"
            "I'm reaching out because I believe there's a fantastic opportunity that aligns "
            "perfectly with your interests and expertise. We'd love to explore how we can work "
            "together and create something meaningful.\n\n"
            "Would you be open to a brief call this week to discuss further? "
            "I'd be happy to work around your schedule.\n\n"
            "Looking forward to connecting!\n\n"
            "Best regards,\n[Your Name]\n[Your Title]\n[Your Company]",
            "data/webinar.png"   # ← put your image filename here, or leave blank
        ),
        (
            "Template 2",
            "Following Up – Quick Question for You",
            "Hi {{Name}},\n\n"
            "I wanted to quickly follow up on my previous message and see if you had a chance to review it.\n\n"
            "I completely understand how busy things can get, so I'll keep this brief — "
            "I genuinely think this could be valuable for you, and I'd love just 15 minutes of your time.\n\n"
            "Please let me know a time that works best for you.\n\n"
            "Warm regards,\n[Your Name]\n[Your Title]\n[Your Company]",
            ""   # ← blank = no image for this template
        ),
        (
            "Template 3",
            "Last Chance – Special Offer Ending Soon",
            "Dear {{Name}},\n\n"
            "This is a friendly reminder that our special offer is available for a limited time only.\n\n"
            "We've helped many professionals like yourself achieve great results, "
            "and we'd hate for you to miss out on this opportunity.\n\n"
            "If you have any questions or would like more information, please don't hesitate to reach out.\n\n"
            "Best wishes,\n[Your Name]\n[Your Title]\n[Your Company]",
            ""   # ← blank = no image
        ),
    ]

    for r, (tno, subject, body, image) in enumerate(templates, 2):
        alt = (r % 2 == 0)
        for c, val in enumerate([tno, subject, body, image], 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data(cell, alt)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
        ws.row_dimensions[r].height = 100

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 28   # Image Path column
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def main():
    os.makedirs(os.path.dirname(Config.EXCEL_FILE), exist_ok=True)

    wb = Workbook()
    build_sheet1(wb)
    build_sheet2(wb)
    wb.save(Config.EXCEL_FILE)

    print(f"✅ Excel file created: {Config.EXCEL_FILE}")
    print("   → Sheet1: Contact list (Name, Email, Email Status, Remark)")
    print("   → Sheet2: 3 email templates ready to use")
    print("\n   Fill in your real contacts in Sheet1, then run: python main.py")


if __name__ == "__main__":
    main()
